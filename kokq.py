import openpyxl
import datetime
import urllib.request as urllib2


def main():
    urllib2.urlretrieve(
        "https://www.vsu.by/images/rasp/mf/2021.05.07/%D0%A4%D0%9C%D0%B8%D0%98%D0%A2_%D1%80%D0%B0%D1%81%D0%BF%D0%B8%D1%81%D0%B0%D0%BD%D0%B8%D0%B5_%D0%B7%D0%B0%D0%BD%D1%8F%D1%82%D0%B8%D0%B9_14_%D0%BD%D0%B5%D0%B4_%D1%81_10.05.2021.xlsx",
        'расписание.xlsx')
    item = open('расписание.xlsx', 'rb')
    wb = openpyxl.load_workbook('расписание.xlsx')
    # РїРѕР»СѓС‡Р°РµРј Р°РєС‚РёРІРЅС‹Р№ Р»РёСЃС‚
    sheet = wb.active

    now = datetime.datetime.now()


    sheet = wb.active
    max_col = sheet.max_column
    day = datetime.datetime.today().weekday()
    # a = ['Понедельник', 'Вторник', 'Среда', 'Четверг',
    #      'Пятникца', 'Суббота', 'Вскр']
    a = ['Понедельник', 'Понедельник', 'Среда', 'Четверг',
         'Пятникца', 'Суббота', 'Среда']

    if now.hour <= 8 or now.hour == 9 & now.minute <= 25:
        para = 1

    if now.hour == 9 & now.hour > 25 & now.hour <= 11:
        para = 2

    if now.hour == 11 & now.minute > 0 or now.hour == 12 & now.minute <= 55:
        para = 3

    if now.hour == 13 & now.minute > 0 or now.hour == 14 & now.minute <= 30:
        para = 4

    if now.hour == 14 & now.minute > 30 or now.hour >= 15 & now.minute <= 55:
        para = 5

    group = 34
    max_row = sheet.max_row
    max_col = sheet.max_column
    R = []

    para = 1;#delete


    # РёС‰РµРј РїРµРІСЂРѕРµ РІС…РѕР¶РґРµРЅРёРµ РґРЅСЏ РЅРµРґРµР»Рё РІ 1РѕРј СЃС‚РѕР»Р±С†Рµ
    for r in range(1, max_row + 1):
        if sheet.cell(row=r, column=1).value == a[day]:
            R.append(r)  # 0 - РґРµРЅСЊ РЅРµРґРµР»Рё
            break

    #######################

    #    R.append(16)
    #    para = 2
    ########################

    # РёС‰РµРј РіСЂСѓРїРїСѓ РїРѕ РєРѕР»РѕРЅРєР°Рј
    for c in range(3, max_col + 1):
        if sheet.cell(row=14, column=c).value == group:
            R.append(c)  # 1 - РєРѕР»РѕРЅРєР° РіСЂСѓРїРїС‹
            break

    # РёС‰РµРј РїР°СЂСѓ РІ 3РµР№ РєРѕР»РѕРЅРєРµ
    for r in range(R[0], R[0] + 14):
        if str(sheet.cell(row=r, column=3).value)[0:1] == str(para):
            R.append(r)  # 2- РЅРѕРІРµСЂ РїР°СЂС‹
            break
    R[2] += 2

    #############
    #    print(str(para) + "- para")
    #    print(str(R[0]) + "- r0")
    #    print(str(R[1]) + "- r01")
    #    print(str(R[2]) + "- r02")
    ############

    if sheet.cell(row=R[2], column=R[1]).value == sheet.cell(row=R[2], column=R[1] + 1).value:
        return "У всей группы аудитория: " + str(
            sheet.cell(row=R[2], column=R[1]).value) + '\n' + "дисициплина: " + str(
            sheet.cell(row=R[2] - 2, column=R[1]).value)
    else:
        return "у 34_1 аудитория: " + str(
            sheet.cell(row=R[2], column=R[1]).value) + '\n' + "дисициплина: " + str(
            sheet.cell(row=R[2] - 2, column=R[
                1]).value) + '\n' + "у 34_2 аудитория: " + str(
            sheet.cell(row=R[2], column=R[1] + 1).value) + '\n' + "дисициплина: " + str(
            sheet.cell(row=R[2] - 2, column=R[1] + 1).value)


    item.close()